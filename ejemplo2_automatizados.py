import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

# 1️⃣ 📂 Cargar el archivo de Excel existente o crear uno nuevo
archivo_excel = "./datos_limpios.xlsx"
archivo_excel_modificado = "./datos_limpios_con_" + "graficas" + ".xlsx" 
hoja_datos = "Sheet1"
hoja_estadisticas = "Estadísticas"

try:
    wb = load_workbook(archivo_excel)  # Cargar si ya existe
    ws_datos = wb[hoja_datos]
except FileNotFoundError:
    from openpyxl import Workbook
    wb = Workbook()
    ws_datos = wb.active
    ws_datos.title = hoja_datos

# 2️⃣ 📊 Leer los datos desde la hoja de Excel
df = pd.read_excel(archivo_excel, sheet_name=hoja_datos)

# Convertir la columna de fecha si es necesario
if "Fecha" in df.columns:
    df["Fecha"] = pd.to_datetime(df["Fecha"])

# 3️⃣ 🔄 Crear una tabla pivote (Ventas por Categoría y Fecha)
pivot_df = df.pivot_table(values='Horas_Capacitacion', columns='Departamento', aggfunc='sum', fill_value=0)

# 4️⃣ 📈 Generar gráficos

# Gráfico de líneas
plt.figure(figsize=(12, 8))
sns.lineplot(data=pivot_df, marker='o')
plt.title('Resultado de capacitacion')
plt.xlabel('Productividad_%')
plt.ylabel('Evaluacion_Desempeno_Despues')
plt.xticks(rotation=45)
plt.legend(title='Categoría')
plt.tight_layout()
plt.savefig("grafico_lineas.png")  # Guardar imagen

# Gráfico de barras
pivot_df.plot(kind='bar', figsize=(12, 8))
plt.title('Horas de capacitacion')
plt.ylabel('Horas de capacitacion')
plt.xlabel('Departamento')
plt.xticks(rotation=45)
plt.legend(title='Categoría')
plt.tight_layout()
plt.savefig("grafico_barras.png")  # Guardar imagen
plt.close()

# 5️⃣ 📊 Calcular estadísticas descriptivas
estadisticas = df.describe()

# 6️⃣ 🔗 Calcular correlaciones
df.columns = df.columns.str.strip()
df = df.reset_index()
correlaciones = df[['Evaluacion_Desempeno_Antes', 'Evaluacion_Desempeno_Despues']].corr()

# 7️⃣ 📂 Guardar estadísticas y correlaciones en Excel

# Crear o acceder a la hoja de estadísticas
if hoja_estadisticas in wb.sheetnames:
    ws_estadisticas = wb[hoja_estadisticas]
    ws_estadisticas.delete_rows(1, ws_estadisticas.max_row)  # Borrar datos antiguos
else:
    ws_estadisticas = wb.create_sheet(title=hoja_estadisticas)

# Escribir estadísticas descriptivas
ws_estadisticas.append(["Estadísticas Descriptivas"])
for row in dataframe_to_rows(estadisticas, index=True, header=True):
    ws_estadisticas.append(row)

# Escribir correlaciones
ws_estadisticas.append([])
ws_estadisticas.append(["Matriz de Correlación"])
for row in dataframe_to_rows(correlaciones, index=True, header=True):
    ws_estadisticas.append(row)

# 8️⃣ 🖼 Insertar imágenes en Excel
ws_datos.add_image(Image("grafico_lineas.png"), "E5")
ws_datos.add_image(Image("grafico_barras.png"), "E20")

# 9️⃣ 📁 Guardar el archivo Excel sin perder datos previos
wb.save(archivo_excel_modificado)

print(f"✅ Reporte generado y actualizado en: {archivo_excel_modificado}")
